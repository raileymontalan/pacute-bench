#!/usr/bin/env python3
"""
Generate Excel annotation workbooks for the human baseline study.

Run from the repo root after sample_human_baseline.py:
    python scripts/generate_annotation_sheets.py [--annotators 3]

One identical workbook per annotator:
    data/human_baseline/annotation_annotator1.xlsx
    data/human_baseline/annotation_annotator2.xlsx
    data/human_baseline/annotation_annotator3.xlsx

Each workbook has one tab per benchmark×format.
MCQ tabs:  item_id | category | subcategory | question_en | question_tl |
           system_prompt (hidden) | option_A..D | correct_option (hidden) |
           annotator_answer (dropdown) | notes
GEN tabs:  item_id | category | subcategory | question_en | question_tl |
           system_prompt | reference_answer (hidden) | annotator_answer | notes
"""
import argparse
import json
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUT_DIR = Path("data/human_baseline")

# Sheet names (≤31 chars)
SHEET_ORDER = [
    ("composition",              "mcq"),
    ("composition",              "gen"),
    ("manipulation",             "mcq"),
    ("manipulation",             "gen"),
    ("syllabification",          "mcq"),
    ("syllabification",          "gen"),
    ("morphological_extraction", "mcq"),
    ("morphological_extraction", "gen"),
    ("morphological_production", "mcq"),
    ("morphological_production", "gen"),
    ("hierarchical",             "mcq"),
    ("hierarchical",             "gen"),
    ("langgame",                 "mcq"),
    ("langgame",                 "gen"),
    ("multi_digit_addition",     "mcq"),
    ("multi_digit_addition",     "gen"),
    ("cute",                     "gen"),
]

SHEET_NAMES = {
    ("composition",              "mcq"): "composition_MCQ",
    ("composition",              "gen"): "composition_GEN",
    ("manipulation",             "mcq"): "manipulation_MCQ",
    ("manipulation",             "gen"): "manipulation_GEN",
    ("syllabification",          "mcq"): "syllabification_MCQ",
    ("syllabification",          "gen"): "syllabification_GEN",
    ("morphological_extraction", "mcq"): "morph_extract_MCQ",
    ("morphological_extraction", "gen"): "morph_extract_GEN",
    ("morphological_production", "mcq"): "morph_prod_MCQ",
    ("morphological_production", "gen"): "morph_prod_GEN",
    ("hierarchical",             "mcq"): "hierarchical_MCQ",
    ("hierarchical",             "gen"): "hierarchical_GEN",
    ("langgame",                 "mcq"): "langgame_MCQ",
    ("langgame",                 "gen"): "langgame_GEN",
    ("multi_digit_addition",     "mcq"): "addition_MCQ",
    ("multi_digit_addition",     "gen"): "addition_GEN",
    ("cute",                     "gen"): "cute_GEN",
}

HEADER_FILL = PatternFill("solid", fgColor="BDD7EE")   # light blue
HIDDEN_FILL = PatternFill("solid", fgColor="D9D9D9")   # grey
ANSWER_FILL = PatternFill("solid", fgColor="E2EFDA")   # light green

MCQ_INSTRUCTION = (
    "Read the question carefully. Select the correct option: A, B, C, or D.\n"
    "Enter your answer in the 'annotator_answer' column only."
)


def _header_font():
    return Font(bold=True)


def _load_items(fmt):
    path = OUT_DIR / f"sample_{fmt}.jsonl"
    items = []
    with open(path, encoding="utf-8") as f:
        for line in f:
            if line.strip():
                items.append(json.loads(line))
    return items


def _group_items(items):
    """Group items by (benchmark, format)."""
    groups = defaultdict(list)
    for item in items:
        groups[(item["benchmark"], item["format"])].append(item)
    return groups


def _col(idx):
    """1-indexed column index → letter."""
    return get_column_letter(idx)


def _write_mcq_sheet(ws, items):
    # Columns:
    # 1:item_id 2:category 3:subcategory 4:question_en 5:question_tl
    # 6:system_prompt(hidden) 7:option_A 8:option_B 9:option_C 10:option_D
    # 11:correct_option(hidden) 12:annotator_answer 13:notes
    headers = [
        "item_id", "category", "subcategory",
        "question_en", "question_tl",
        "system_prompt",
        "option_A", "option_B", "option_C", "option_D",
        "correct_option",
        "annotator_answer", "notes",
    ]
    hidden_cols = {6, 11}   # system_prompt, correct_option
    answer_col = 12

    # Write header
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = _header_font()
        cell.fill = HIDDEN_FILL if c in hidden_cols else (
            ANSWER_FILL if c == answer_col else HEADER_FILL
        )
        cell.alignment = Alignment(wrap_text=True)

    ws.freeze_panes = "A2"

    # Column widths
    ws.column_dimensions["A"].width = 28       # item_id
    ws.column_dimensions["B"].width = 20       # category
    ws.column_dimensions["C"].width = 22       # subcategory
    ws.column_dimensions["D"].width = 55       # question_en
    ws.column_dimensions["E"].width = 55       # question_tl
    ws.column_dimensions["F"].width = 0.1      # system_prompt (hidden)
    ws.column_dimensions["G"].width = 25       # option_A
    ws.column_dimensions["H"].width = 25       # option_B
    ws.column_dimensions["I"].width = 25       # option_C
    ws.column_dimensions["J"].width = 25       # option_D
    ws.column_dimensions["K"].width = 0.1      # correct_option (hidden)
    ws.column_dimensions["L"].width = 20       # annotator_answer
    ws.column_dimensions["M"].width = 30       # notes

    ws.column_dimensions["F"].hidden = True
    ws.column_dimensions["K"].hidden = True

    # Write data + dropdowns
    for row_idx, item in enumerate(items, 2):
        opts = item.get("options", {})
        n_opts = item.get("n_options", len(opts))
        valid_letters = ",".join(list(opts.keys())[:n_opts])

        dv = DataValidation(
            type="list",
            formula1=f'"{valid_letters}"',
            allow_blank=True,
            showErrorMessage=True,
            error="Enter one of: " + valid_letters,
            errorTitle="Invalid answer",
        )
        ws.add_data_validation(dv)
        answer_cell = ws.cell(row=row_idx, column=answer_col)
        dv.add(answer_cell)
        answer_cell.fill = ANSWER_FILL

        ws.cell(row=row_idx, column=1,  value=item["item_id"])
        ws.cell(row=row_idx, column=2,  value=item.get("category") or "")
        ws.cell(row=row_idx, column=3,  value=item.get("subcategory") or "")
        ws.cell(row=row_idx, column=4,  value=item.get("question_en") or "")
        ws.cell(row=row_idx, column=5,  value=item.get("question_tl") or "")
        ws.cell(row=row_idx, column=6,  value=MCQ_INSTRUCTION)
        ws.cell(row=row_idx, column=7,  value=opts.get("A", ""))
        ws.cell(row=row_idx, column=8,  value=opts.get("B", ""))
        ws.cell(row=row_idx, column=9,  value=opts.get("C", ""))
        ws.cell(row=row_idx, column=10, value=opts.get("D", ""))
        ws.cell(row=row_idx, column=11, value=item.get("correct_option") or "")
        ws.cell(row=row_idx, column=13, value="")

        for c in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=c).alignment = Alignment(wrap_text=True, vertical="top")


def _write_gen_sheet(ws, items):
    # Columns:
    # 1:item_id 2:category 3:subcategory 4:question_en 5:question_tl
    # 6:system_prompt 7:reference_answer(hidden) 8:annotator_answer 9:notes
    headers = [
        "item_id", "category", "subcategory",
        "question_en", "question_tl",
        "system_prompt",
        "reference_answer",
        "annotator_answer", "notes",
    ]
    hidden_cols = {7}
    answer_col = 8

    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = _header_font()
        cell.fill = HIDDEN_FILL if c in hidden_cols else (
            ANSWER_FILL if c == answer_col else HEADER_FILL
        )
        cell.alignment = Alignment(wrap_text=True)

    ws.freeze_panes = "A2"

    ws.column_dimensions["A"].width = 32       # item_id
    ws.column_dimensions["B"].width = 20       # category
    ws.column_dimensions["C"].width = 22       # subcategory
    ws.column_dimensions["D"].width = 60       # question_en
    ws.column_dimensions["E"].width = 60       # question_tl
    ws.column_dimensions["F"].width = 50       # system_prompt
    ws.column_dimensions["G"].width = 0.1      # reference_answer (hidden)
    ws.column_dimensions["H"].width = 30       # annotator_answer
    ws.column_dimensions["I"].width = 30       # notes

    ws.column_dimensions["G"].hidden = True

    for row_idx, item in enumerate(items, 2):
        ws.cell(row=row_idx, column=1, value=item["item_id"])
        ws.cell(row=row_idx, column=2, value=item.get("category") or "")
        ws.cell(row=row_idx, column=3, value=item.get("subcategory") or "")
        ws.cell(row=row_idx, column=4, value=item.get("question_en") or "")
        ws.cell(row=row_idx, column=5, value=item.get("question_tl") or "")
        ws.cell(row=row_idx, column=6, value=item.get("system_prompt") or "")
        ws.cell(row=row_idx, column=7, value=item.get("reference_answer") or "")
        ws.cell(row=row_idx, column=8, value="")
        ws.cell(row=row_idx, column=9, value="")

        ws.cell(row=row_idx, column=8).fill = ANSWER_FILL

        for c in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=c).alignment = Alignment(wrap_text=True, vertical="top")

        # Row height: taller rows for long system prompts
        ws.row_dimensions[row_idx].height = 60


def build_workbook(mcq_groups, gen_groups, annotator_id):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    # Add a cover sheet
    cover = wb.create_sheet("README")
    cover["A1"] = f"PACUTE-Bench Human Baseline — Annotator {annotator_id}"
    cover["A1"].font = Font(bold=True, size=14)
    cover["A3"] = "Instructions:"
    cover["A3"].font = Font(bold=True)
    cover["A4"] = (
        "1. Work through each tab independently.\n"
        "2. For MCQ tabs: select A, B, C, or D from the dropdown in the 'annotator_answer' column.\n"
        "3. For GEN tabs: type your free-form answer in the 'annotator_answer' column.\n"
        "   Follow the format described in the 'system_prompt' column.\n"
        "4. Do not leave 'annotator_answer' blank — blank = incorrect.\n"
        "5. Use the 'notes' column if unsure; do NOT look up answers.\n"
        "6. Hidden columns (grey headers) are for scoring only — do not unhide/edit them."
    )
    cover["A4"].alignment = Alignment(wrap_text=True, vertical="top")
    cover.column_dimensions["A"].width = 80
    cover.row_dimensions[4].height = 100

    all_groups = {**mcq_groups, **gen_groups}

    for bm, fmt in SHEET_ORDER:
        key = (bm, fmt)
        items = all_groups.get(key, [])
        if not items:
            continue

        ws = wb.create_sheet(SHEET_NAMES[key])
        if fmt == "mcq":
            _write_mcq_sheet(ws, items)
        else:
            _write_gen_sheet(ws, items)

        ws.sheet_properties.tabColor = "BDD7EE" if fmt == "mcq" else "E2EFDA"

    return wb


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--annotators", type=int, default=3)
    args = parser.parse_args()

    mcq_items = _load_items("mcq")
    gen_items = _load_items("gen")

    mcq_groups = _group_items(mcq_items)
    gen_groups = _group_items(gen_items)

    # Summary
    total_mcq = sum(len(v) for v in mcq_groups.values())
    total_gen = sum(len(v) for v in gen_groups.values())
    print(f"Loaded {total_mcq} MCQ items, {total_gen} GEN items → {total_mcq + total_gen} total per annotator")

    OUT_DIR.mkdir(parents=True, exist_ok=True)

    for i in range(1, args.annotators + 1):
        path = OUT_DIR / f"annotation_annotator{i}.xlsx"
        wb = build_workbook(mcq_groups, gen_groups, annotator_id=i)
        wb.save(path)
        print(f"Written → {path}")


if __name__ == "__main__":
    main()
