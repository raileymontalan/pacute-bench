#!/usr/bin/env python3
"""
Score filled human baseline annotation workbooks and compute IAA.

Run from the repo root:
    python scripts/score_human_baselines.py \
        data/human_baseline_annotations/annotation_annotator1.xlsx \
        data/human_baseline_annotations/annotation_annotator2.xlsx \
        data/human_baseline_annotations/annotation_annotator3.xlsx

Outputs (written to results/human_baseline/):
    scores_per_annotator.json   — per-annotator accuracy/exact_match by benchmark
    iaa.json                    — Fleiss' kappa and percent agreement
    summary.md                  — Markdown table for the paper
"""
import argparse
import json
import math
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl

OUT_DIR = Path("results-202605/human_baseline")
SAMPLE_DIR = Path("data/human_baseline_annotations")

# Sheet name → (benchmark, format) reverse map
SHEET_TO_KEY = {v: k for k, v in {
    ("composition",              "mcq"): "composition_MCQ",
    ("composition",              "gen"): "composition_GEN",
    ("manipulation",             "mcq"): "manipulation_MCQ",
    ("manipulation",             "gen"): "manipulation_GEN",
    ("syllabification",          "mcq"): "syllabification_MCQ",
    ("syllabification",          "gen"): "syllabification_GEN",
    ("morphological_extraction", "mcq"): "morph_extract_MCQ",
    ("morphological_extraction", "gen"): "morph_extract_GEN",
    ("morphological_production", "mcq"): "morph_prod_MCQ",
    ("morphological_production", "gen"): "morph_prod_GEN",
    ("hierarchical",             "mcq"): "hierarchical_MCQ",
    ("hierarchical",             "gen"): "hierarchical_GEN",
    ("langgame",                 "mcq"): "langgame_MCQ",
    ("langgame",                 "gen"): "langgame_GEN",
    ("multi_digit_addition",     "mcq"): "addition_MCQ",
    ("multi_digit_addition",     "gen"): "addition_GEN",
    ("cute",                     "gen"): "cute_GEN",
}.items()}


# --- Normalization (mirrors BaseEvaluator._normalize_label) ------------------

def _normalize(s):
    return str(s).strip().lower().strip("-")


# --- Scoring -----------------------------------------------------------------

def score_mcq(annotator_answer, correct_option):
    """Return 1 if correct, 0 otherwise."""
    if not annotator_answer:
        return 0
    return int(str(annotator_answer).strip().upper() == str(correct_option).strip().upper())


def score_gen(annotator_answer, reference_answer):
    """Return dict with exact_match, contains_match, prefix_match."""
    pred = _normalize(annotator_answer)
    ref = _normalize(reference_answer)
    return {
        "exact_match":    int(pred == ref),
        "contains_match": int(ref in pred),
        "prefix_match":   int(pred.startswith(ref)),
    }


# --- Fleiss' kappa -----------------------------------------------------------

def fleiss_kappa(ratings):
    """
    Compute Fleiss' kappa for N items × n annotators.

    ratings: list of lists, each inner list contains one rating per annotator.
             Ratings can be strings (e.g. "A","B") or ints (0/1).
    Returns: float kappa, or None if undefined (e.g. all same category).
    """
    N = len(ratings)
    if N == 0:
        return None
    n = len(ratings[0])
    if n < 2:
        return None

    categories = sorted({r for row in ratings for r in row})
    k = len(categories)
    if k <= 1:
        return 1.0  # perfect agreement by definition

    cat_idx = {c: i for i, c in enumerate(categories)}

    P_i_list = []
    cat_totals = [0] * k

    for row in ratings:
        counts = [0] * k
        for r in row:
            counts[cat_idx[r]] += 1
            cat_totals[cat_idx[r]] += 1
        p_i = (sum(c * c for c in counts) - n) / (n * (n - 1))
        P_i_list.append(p_i)

    P_bar = sum(P_i_list) / N
    P_e = sum((t / (N * n)) ** 2 for t in cat_totals)

    if abs(1 - P_e) < 1e-10:
        return None  # degenerate
    return (P_bar - P_e) / (1 - P_e)


def percent_exact_agreement(ratings):
    """Fraction of items where all annotators gave the identical answer."""
    if not ratings:
        return 0.0
    return sum(1 for row in ratings if len(set(row)) == 1) / len(ratings)


# --- Workbook reading --------------------------------------------------------

def _read_sheet(ws, fmt):
    """
    Read all rows from an annotation sheet.

    Returns list of dicts: {item_id, annotator_answer, correct_option_or_ref, category, subcategory}
    """
    headers = {cell.value: cell.column for cell in ws[1] if cell.value}

    required = ["item_id", "annotator_answer"]
    for col in required:
        if col not in headers:
            print(f"  WARNING: column '{col}' not found in sheet '{ws.title}'", file=sys.stderr)
            return []

    ref_col = "correct_option" if fmt == "mcq" else "reference_answer"
    if ref_col not in headers:
        print(f"  WARNING: column '{ref_col}' not found in sheet '{ws.title}'", file=sys.stderr)
        return []

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        def get(col):
            idx = headers.get(col)
            if idx is None:
                return None
            v = row[idx - 1]
            return v

        item_id  = get("item_id")
        answer   = get("annotator_answer")
        ref      = get(ref_col)
        category = get("category")
        subcat   = get("subcategory")

        if item_id is None:
            continue

        rows.append({
            "item_id":    str(item_id),
            "answer":     str(answer).strip() if answer is not None else "",
            "reference":  str(ref).strip() if ref is not None else "",
            "category":   str(category) if category else None,
            "subcategory": str(subcat) if subcat else None,
        })
    return rows


def read_workbook(path):
    """Return dict: sheet_name → list of row dicts."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    result = {}
    for sheet_name in wb.sheetnames:
        if sheet_name == "README":
            continue
        key = SHEET_TO_KEY.get(sheet_name)
        if key is None:
            continue
        bm, fmt = key
        ws = wb[sheet_name]
        rows = _read_sheet(ws, fmt)
        result[key] = rows
    wb.close()
    return result


# --- Per-annotator scoring ---------------------------------------------------

def score_annotator(data_by_key):
    """
    data_by_key: dict (benchmark, format) → list of row dicts
    Returns nested dict: {(bm, fmt): {overall: {...}, by_category: {...}}}
    """
    results = {}
    for key, rows in data_by_key.items():
        bm, fmt = key
        if fmt == "mcq":
            results[key] = _score_mcq_rows(rows)
        else:
            results[key] = _score_gen_rows(rows)
    return results


def _score_mcq_rows(rows):
    items = []
    for r in rows:
        correct = score_mcq(r["answer"], r["reference"])
        items.append({"item_id": r["item_id"], "category": r["category"],
                       "subcategory": r["subcategory"], "is_correct": correct,
                       "answer": r["answer"], "reference": r["reference"]})
    return _aggregate_mcq(items)


def _aggregate_mcq(items):
    n = len(items)
    if n == 0:
        return {"num_samples": 0}
    correct = sum(i["is_correct"] for i in items)
    acc = correct / n

    # Build by_category
    by_cat = defaultdict(list)
    for i in items:
        cat = (i["category"] or "") + (f"/{i['subcategory']}" if i["subcategory"] else "")
        by_cat[cat or "__all__"].append(i)

    by_category = {}
    for cat, cat_items in sorted(by_cat.items()):
        nc = len(cat_items)
        cc = sum(i["is_correct"] for i in cat_items)
        by_category[cat] = {
            "num_samples": nc,
            "accuracy": round(cc / nc, 4),
        }

    return {
        "num_samples": n,
        "accuracy": round(acc, 4),
        "by_category": by_category,
    }


def _score_gen_rows(rows):
    items = []
    for r in rows:
        scores = score_gen(r["answer"], r["reference"])
        items.append({"item_id": r["item_id"], "category": r["category"],
                       "subcategory": r["subcategory"], "answer": r["answer"],
                       "reference": r["reference"], **scores})
    return _aggregate_gen(items)


def _aggregate_gen(items):
    n = len(items)
    if n == 0:
        return {"num_samples": 0}

    by_cat = defaultdict(list)
    for i in items:
        cat = (i["category"] or "") + (f"/{i['subcategory']}" if i["subcategory"] else "")
        by_cat[cat or "__all__"].append(i)

    by_category = {}
    for cat, cat_items in sorted(by_cat.items()):
        nc = len(cat_items)
        by_category[cat] = {
            "num_samples": nc,
            "exact_match":    round(sum(i["exact_match"]    for i in cat_items) / nc, 4),
            "contains_match": round(sum(i["contains_match"] for i in cat_items) / nc, 4),
            "prefix_match":   round(sum(i["prefix_match"]   for i in cat_items) / nc, 4),
        }

    return {
        "num_samples":    n,
        "exact_match":    round(sum(i["exact_match"]    for i in items) / n, 4),
        "contains_match": round(sum(i["contains_match"] for i in items) / n, 4),
        "prefix_match":   round(sum(i["prefix_match"]   for i in items) / n, 4),
        "by_category": by_category,
    }


# --- IAA ---------------------------------------------------------------------

def compute_iaa(annotator_data_list):
    """
    annotator_data_list: list of dicts from read_workbook(), one per annotator.
    Returns: {(bm, fmt): {"kappa": float, "pct_agree": float, "n": int, ...}}
    """
    # Find all keys present in all annotators
    all_keys = set(annotator_data_list[0].keys())
    for ad in annotator_data_list[1:]:
        all_keys &= set(ad.keys())

    iaa = {}
    for key in sorted(all_keys):
        bm, fmt = key

        # Build item_id → list of answers (one per annotator)
        item_answers = defaultdict(list)
        item_refs = {}
        for ad in annotator_data_list:
            rows = ad.get(key, [])
            for r in rows:
                item_answers[r["item_id"]].append(r["answer"])
                item_refs[r["item_id"]] = r["reference"]

        # Only items answered by ALL annotators
        n_ann = len(annotator_data_list)
        complete = {iid: ans for iid, ans in item_answers.items() if len(ans) == n_ann}

        if not complete:
            iaa[key] = {"n": 0}
            continue

        ratings = list(complete.values())

        if fmt == "mcq":
            kappa = fleiss_kappa(ratings)
            pct   = percent_exact_agreement(ratings)
            iaa[key] = {
                "n":        len(ratings),
                "kappa":    round(kappa, 4) if kappa is not None else None,
                "pct_agree": round(pct, 4),
            }
        else:
            # Binary kappa: correct (1) vs incorrect (0) per annotator
            binary_ratings = []
            for iid, ans_list in complete.items():
                ref = item_refs[iid]
                binary_ratings.append([score_mcq(a, _normalize(ref)) for a in ans_list])

            # Also raw answer agreement (all annotators gave same answer)
            kappa_binary = fleiss_kappa(binary_ratings)
            pct_raw      = percent_exact_agreement(ratings)
            pct_binary   = percent_exact_agreement(binary_ratings)
            iaa[key] = {
                "n":                   len(ratings),
                "kappa_binary":        round(kappa_binary, 4) if kappa_binary is not None else None,
                "pct_exact_agree_raw": round(pct_raw, 4),
                "pct_exact_agree_bin": round(pct_binary, 4),
            }

    return iaa


# --- Summary markdown ---------------------------------------------------------

def build_summary(per_annotator_scores, iaa):
    """Build a Markdown summary table."""
    lines = ["# Human Baseline Results\n"]

    # Collect all keys
    all_keys = sorted({k for ann_scores in per_annotator_scores for k in ann_scores})

    lines.append("## MCQ Benchmarks\n")
    lines.append("| Benchmark | Ann1 acc | Ann2 acc | Ann3 acc | Mean±Std | Fleiss κ | % agree |")
    lines.append("|---|---|---|---|---|---|---|")
    for key in all_keys:
        bm, fmt = key
        if fmt != "mcq":
            continue
        accs = [
            ann_scores[key]["accuracy"]
            for ann_scores in per_annotator_scores
            if key in ann_scores and "accuracy" in ann_scores[key]
        ]
        if not accs:
            continue
        mean = sum(accs) / len(accs)
        std  = math.sqrt(sum((a - mean)**2 for a in accs) / len(accs))
        k_info = iaa.get(key, {})
        kappa  = k_info.get("kappa", "N/A")
        pct    = k_info.get("pct_agree", "N/A")
        kappa_str = f"{kappa:.3f}" if isinstance(kappa, float) else str(kappa)
        pct_str   = f"{float(pct)*100:.1f}%" if isinstance(pct, float) else str(pct)
        ann_cols  = " | ".join(f"{a:.3f}" for a in accs)
        lines.append(f"| {bm} | {ann_cols} | {mean:.3f}±{std:.3f} | {kappa_str} | {pct_str} |")

    lines.append("\n## GEN Benchmarks\n")
    lines.append("| Benchmark | Ann1 EM | Ann2 EM | Ann3 EM | Mean±Std | κ (binary) | raw agree% |")
    lines.append("|---|---|---|---|---|---|---|")
    for key in all_keys:
        bm, fmt = key
        if fmt != "gen":
            continue
        ems = [
            ann_scores[key]["exact_match"]
            for ann_scores in per_annotator_scores
            if key in ann_scores and "exact_match" in ann_scores[key]
        ]
        if not ems:
            continue
        mean = sum(ems) / len(ems)
        std  = math.sqrt(sum((e - mean)**2 for e in ems) / len(ems))
        k_info = iaa.get(key, {})
        kappa  = k_info.get("kappa_binary", "N/A")
        pct    = k_info.get("pct_exact_agree_raw", "N/A")
        kappa_str = f"{kappa:.3f}" if isinstance(kappa, float) else str(kappa)
        pct_str   = f"{float(pct)*100:.1f}%" if isinstance(pct, float) else str(pct)
        ann_cols  = " | ".join(f"{e:.3f}" for e in ems)
        lines.append(f"| {bm} | {ann_cols} | {mean:.3f}±{std:.3f} | {kappa_str} | {pct_str} |")

    return "\n".join(lines)


# --- Main --------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("workbooks", nargs="+", help="Filled annotation workbook paths")
    args = parser.parse_args()

    if len(args.workbooks) < 2:
        print("Need at least 2 workbooks for IAA.", file=sys.stderr)

    annotator_raw = []
    for path in args.workbooks:
        print(f"Reading {path}...")
        data = read_workbook(Path(path))
        annotator_raw.append(data)
        for key, rows in data.items():
            bm, fmt = key
            answered = sum(1 for r in rows if r["answer"])
            print(f"  {bm} {fmt}: {answered}/{len(rows)} answered")

    print("\nScoring per annotator...")
    per_annotator_scores = [score_annotator(data) for data in annotator_raw]

    print("Computing IAA...")
    iaa = compute_iaa(annotator_raw)

    OUT_DIR.mkdir(parents=True, exist_ok=True)

    # Serialize keys as strings for JSON
    def key_str(k):
        return f"{k[0]}_{k[1]}"

    scores_out = {
        f"annotator{i+1}": {key_str(k): v for k, v in scores.items()}
        for i, scores in enumerate(per_annotator_scores)
    }
    iaa_out = {key_str(k): v for k, v in iaa.items()}

    with open(OUT_DIR / "scores_per_annotator.json", "w", encoding="utf-8") as f:
        json.dump(scores_out, f, indent=2, ensure_ascii=False)
    print(f"Written → {OUT_DIR}/scores_per_annotator.json")

    with open(OUT_DIR / "iaa.json", "w", encoding="utf-8") as f:
        json.dump(iaa_out, f, indent=2, ensure_ascii=False)
    print(f"Written → {OUT_DIR}/iaa.json")

    summary_md = build_summary(per_annotator_scores, iaa)
    with open(OUT_DIR / "summary.md", "w", encoding="utf-8") as f:
        f.write(summary_md)
    print(f"Written → {OUT_DIR}/summary.md")

    # Quick console summary
    print("\n=== Quick IAA Summary ===")
    for key, info in sorted(iaa.items()):
        bm, fmt = key
        n = info.get("n", 0)
        if fmt == "mcq":
            k = info.get("kappa", "N/A")
            p = info.get("pct_agree", "N/A")
            k_str = f"{k:.3f}" if isinstance(k, float) else str(k)
            p_str = f"{float(p)*100:.1f}%" if isinstance(p, float) else str(p)
            print(f"  {bm} MCQ  (n={n}): κ={k_str}  agree={p_str}")
        else:
            k = info.get("kappa_binary", "N/A")
            p = info.get("pct_exact_agree_raw", "N/A")
            k_str = f"{k:.3f}" if isinstance(k, float) else str(k)
            p_str = f"{float(p)*100:.1f}%" if isinstance(p, float) else str(p)
            print(f"  {bm} GEN  (n={n}): κ_bin={k_str}  raw_agree={p_str}")


if __name__ == "__main__":
    main()
