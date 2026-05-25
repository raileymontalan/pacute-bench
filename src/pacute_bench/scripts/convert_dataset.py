#!/usr/bin/env python3
"""
Convert pacute_dataset.xlsx into JSONL corpus files.

Outputs:
  data/corpora/pacute_data/corpus_composition.jsonl
  data/corpora/pacute_data/corpus_morphological_extraction.jsonl

Usage:
    python -m pacute_bench.scripts.convert_dataset
    python -m pacute_bench.scripts.convert_dataset --xlsx pacute_dataset.xlsx --output-dir data/corpora/pacute_data
"""
from __future__ import annotations

import argparse
import json
from pathlib import Path

import openpyxl


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _read_sheet(wb: openpyxl.Workbook, sheet_name: str) -> list[dict]:
    """Read a worksheet into a list of dicts (data_only values, skip None-key cols)."""
    ws = wb[sheet_name]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    rows = []
    for r in range(2, ws.max_row + 1):
        row = {}
        for c, h in enumerate(headers, 1):
            if h is not None:
                row[h] = ws.cell(r, c).value
        # Skip entirely empty rows
        if any(v is not None for v in row.values()):
            rows.append(row)
    return rows


def _write_jsonl(rows: list[dict], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        for row in rows:
            f.write(json.dumps(row, ensure_ascii=False) + "\n")
    print(f"  Wrote {len(rows)} rows → {path}")


# ---------------------------------------------------------------------------
# Composition corpus
# ---------------------------------------------------------------------------

def _convert_composition(wb: openpyxl.Workbook) -> list[dict]:
    rows = []

    # diacritics
    for r in _read_sheet(wb, "diacritics"):
        rows.append({
            "category":        r.get("category"),
            "subcategory":     "diacritics",
            "text_en":         r.get("text_en"),
            "text_tl":         r.get("text_tl"),
            "label":           str(r.get("label", "")),
            "id":              r.get("id"),
            "word":            r.get("diacritized_word"),
            "normalized_form": r.get("normalized_form"),
            "diacritics_count": r.get("diacritics_count"),
        })

    # uppercasing
    for r in _read_sheet(wb, "uppercasing"):
        rows.append({
            "category":        r.get("category"),
            "subcategory":     "uppercasing",
            "text_en":         r.get("text_en"),
            "text_tl":         r.get("text_tl"),
            "label":           str(r.get("label", "")),
            "id":              r.get("id"),
            "word":            r.get("word"),
            "normalized_form": r.get("normalized_form"),
            "uppercase_count": r.get("uppercase_count"),
        })

    for r in _read_sheet(wb, "character_counting"):
        rows.append({
            "category":    r.get("category"),
            "subcategory": "character_counting",
            "subcategory2": r.get("subcategory2"),
            "text_en":     r.get("text_en"),
            "text_tl":     r.get("text_tl"),
            "label":       str(r.get("label", "")),
            "id":          r.get("id"),
            "word":        r.get("word"),
        })

    # character_recognition (exclude ng_aware rows)
    for r in _read_sheet(wb, "character_recognition"):
        if r.get("subcategory2") == "ng_aware":
            continue
        rows.append({
            "category":    r.get("category"),
            "subcategory": "character_recognition",
            "subcategory2": r.get("subcategory2"),
            "text_en":     r.get("text_en"),
            "text_tl":     r.get("text_tl"),
            "label":       str(r.get("label", "")),
            "id":          r.get("id"),
            "word":        r.get("word"),
            "position":    r.get("position"),
            "position_english": r.get("position_english"),
        })

    return rows


# ---------------------------------------------------------------------------
# Morphological extraction corpus
# ---------------------------------------------------------------------------

def _convert_morphological_extraction(wb: openpyxl.Workbook) -> list[dict]:
    rows = []

    # inflected_affix_extraction
    for r in _read_sheet(wb, "inflected_affix_extraction"):
        rows.append({
            "category":    r.get("category"),
            "subcategory": "inflected_affix_extraction",
            "subcategory2": r.get("subcategory2"),
            "subcategory3": r.get("subcategory3"),
            "text_en":     r.get("text_en"),
            "text_tl":     r.get("text_tl"),
            "label":       str(r.get("label", "")),
            "id":          r.get("id"),
            "word":        r.get("word"),
            "root":        r.get("root_word"),
            "affix":       r.get("affix"),
            "affix_alternate": r.get("affix_alternate"),
            "affix_type":  r.get("affix_type"),
            "subtype":     r.get("subtype"),
        })

    # inflected_root_extraction
    for r in _read_sheet(wb, "inflected_root_extraction"):
        rows.append({
            "category":    r.get("category"),
            "subcategory": "inflected_root_extraction",
            "subcategory2": r.get("subcategory2"),
            "subcategory3": r.get("subcategory3"),
            "text_en":     r.get("text_en"),
            "text_tl":     r.get("text_tl"),
            "label":       str(r.get("label", "")),
            "id":          r.get("id"),
            "word":        r.get("word"),
            "root":        r.get("root_word"),
            "affix":       r.get("affix"),
            "affix_alternate": r.get("affix_alternate"),
            "affix_type":  r.get("affix_type"),
            "subtype":     r.get("subtype"),
        })

    # reduplicated_root_extraction
    for r in _read_sheet(wb, "reduplicated_root_extraction"):
        rows.append({
            "category":    r.get("category"),
            "subcategory": "reduplicated_root_extraction",
            "subcategory2": r.get("subcategory2"),
            "subcategory3": r.get("subcategory3"),
            "text_en":     r.get("text_en"),
            "text_tl":     r.get("text_tl"),
            "label":       str(r.get("label", "")),
            "id":          r.get("id"),
            "word":        r.get("word"),
            "root":        r.get("root"),
            "reduplicant": r.get("reduplicant"),
            "redup_type":  r.get("redup_type"),
            "subtype":     r.get("subtype"),
        })

    # reduplicant_extraction (sheet name has typo: "reduplicant_extration")
    for r in _read_sheet(wb, "reduplicant_extration"):
        rows.append({
            "category":    r.get("category"),
            "subcategory": "reduplicant_extraction",
            "subcategory2": r.get("subcategory2"),
            "subcategory3": r.get("subcategory3"),
            "text_en":     r.get("text_en"),
            "text_tl":     r.get("text_tl"),
            "label":       str(r.get("label", "")),
            "id":          r.get("id"),
            "word":        r.get("word"),
            "root":        r.get("root"),
            "reduplicant": r.get("reduplicant"),
            "redup_type":  r.get("redup_type"),
            "subtype":     r.get("subtype"),
        })

    return rows


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def convert_pacute_dataset(xlsx_path: Path, output_dir: Path) -> None:
    print(f"Loading {xlsx_path} …")
    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)

    print("\n[1/2] Composition corpus …")
    comp_rows = _convert_composition(wb)
    _write_jsonl(comp_rows, output_dir / "corpus_composition.jsonl")

    print("\n[2/2] Morphological extraction corpus …")
    morph_rows = _convert_morphological_extraction(wb)
    _write_jsonl(morph_rows, output_dir / "corpus_morphological_extraction.jsonl")

    print("\nDone.")


def main() -> None:
    parser = argparse.ArgumentParser(description="Convert pacute_dataset.xlsx to JSONL corpora")
    parser.add_argument("--xlsx", default="pacute_dataset.xlsx",
                        help="Path to pacute_dataset.xlsx (default: pacute_dataset.xlsx)")
    parser.add_argument("--output-dir", default="data/corpora/pacute_data",
                        help="Output directory (default: data/corpora/pacute_data)")
    args = parser.parse_args()
    convert_pacute_dataset(Path(args.xlsx), Path(args.output_dir))


if __name__ == "__main__":
    main()
